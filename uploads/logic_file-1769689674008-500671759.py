"""
conversion_logic_hubx.py
Standardizes HUBX's price-lists.
Logic:
- Input: .xlsx file.
- IGNORES Sheets: "Today's Deals", "Price Drop", "Just Launched".
- PROCESSES all other sheets.
- IGNORES Hidden Rows.
- Fixes line breaks.
- ROUNDS PRICES to 2 decimal places.
- Mapping:
    - Col A = Model
    - Col B = Name
    - Col D = Stock
    - Col E = Price (USD)
    - Notes = Merge Cols J, C, F, G, H, I, K, L
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def standardize(input_path: str, output_path: str):
    print("Starting HUBX conversion...")
    
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    
    # Sheets to explicitly ignore
    IGNORE_SHEETS = [
        "today's deals", 
        "price drop", 
        "just launched"
    ]

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in IGNORE_SHEETS:
            continue

        ws = wb[sheet_name]
        
        # Configuration
        # Header = Row 1, Data starts Row 2
        start_row = 2
        
        # Column Indices (0-based)
        # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11
        idx_model = 0   # Col A
        idx_name = 1    # Col B
        idx_stock = 3   # Col D
        idx_price = 4   # Col E
        
        # Notes merge list: J, C, F, G, H, I, K, L
        # Indices: 9, 2, 5, 6, 7, 8, 10, 11
        idx_notes_list = [9, 2, 5, 6, 7, 8, 10, 11]

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. NAME
            final_name = clean_text(get_val(idx_name))
            if not final_name:
                continue

            # 3. PRICE (USD)
            raw_price = get_val(idx_price)
            clean_price = "0"
            
            try:
                # Remove '$' and ',' then convert
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    val = float(numeric_str)
                    val = round(val, 2) # Round to 2 decimals
                    
                    if val.is_integer():
                        clean_price = str(int(val))
                    else:
                        clean_price = str(val)
            except:
                if "call" in raw_price.lower():
                    clean_price = "CALL"
                else:
                    clean_price = "0"

            # 4. STOCK
            raw_stock = get_val(idx_stock)
            clean_stock = "0"
            try:
                # Sanitize stock (remove non-numeric except digits)
                numeric_stock = "".join(c for c in raw_stock if c.isdigit())
                if numeric_stock:
                    clean_stock = str(int(numeric_stock))
            except:
                clean_stock = "0"

            # 5. MODEL
            final_model = clean_text(get_val(idx_model))

            # 6. NOTES (Merge multiple columns)
            notes_parts = []
            for n_idx in idx_notes_list:
                val = clean_text(get_val(n_idx))
                if val and val.lower() != "nan":
                    notes_parts.append(val)
            
            final_notes = " | ".join(notes_parts)

            products.append({
                "Supplier": "HUBX",
                "Category": clean_text(sheet_name), # Sheet Name = Category
                "Brand": "", # Not specified, blank
                "Model": final_model,
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", 
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Output
    df_out = pd.DataFrame(products)
    
    if df_out.empty:
        print("Warning: No products found for HUBX.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from HUBX.")