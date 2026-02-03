"""
conversion_logic_zotac.py
Standardizes ZOTAC's price-lists.
Logic:
- Input: .xlsx file.
- FINDS Sheet containing "pricelist" (case-insensitive).
- IGNORES Hidden Rows.
- Fixes line breaks.
- ROUNDS PRICES to 2 decimal places.
- Header spans rows 14-15 -> Data starts Row 16.
- Mapping:
    - Brand = "Zotac"
    - Col A = MOQ
    - Col B = Model
    - Col E+F+G+H = Name
    - Col D = Price (USD) - Handles "n/a" by setting to 0 (but keeping product).
    - Col AN = Notes
    - Stock = Default "1" (Not specified in file).
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def standardize(input_path: str, output_path: str):
    print("Starting ZOTAC conversion...")
    
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    
    # 1. FIND RELEVANT SHEET
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "pricelist" in sheet_name.lower():
            target_sheet = wb[sheet_name]
            break
    
    if not target_sheet:
        print("Error: No sheet containing 'pricelist' found.")
        return

    ws = target_sheet
    
    # Configuration
    # Header = Row 14 & 15. Data likely starts Row 16.
    start_row = 16
    
    # Column Indices (0-based)
    # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7 ... AN=39
    idx_moq = 0
    idx_model = 1
    idx_price = 3
    idx_name_parts = [4, 5, 6, 7] # E, F, G, H
    idx_notes = 39 # AN is the 40th column (Index 39)

    for row in ws.iter_rows(min_row=start_row, values_only=False):
        # 2. SKIP HIDDEN ROWS
        try:
            if ws.row_dimensions[row[0].row].hidden:
                continue
        except:
            pass

        cells = [cell.value for cell in row]
        
        def get_val(idx):
            if idx is None: return ""
            if 0 <= idx < len(cells):
                val = cells[idx]
                return str(val).strip() if val is not None else ""
            return ""

        # 3. NAME (Merge E+F+G+H)
        name_components = []
        for i in idx_name_parts:
            val = clean_text(get_val(i))
            if val:
                name_components.append(val)
        final_name = " ".join(name_components)
        
        if not final_name:
            continue

        # 4. PRICE (USD)
        raw_price = get_val(idx_price)
        clean_price = "0"
        price_note = ""
        
        # Handle "n/a" or text
        if raw_price:
            try:
                # Remove '$' and ',' then convert
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    val = float(numeric_str)
                    val = round(val, 2)
                    
                    if val.is_integer():
                        clean_price = str(int(val))
                    else:
                        clean_price = str(val)
                else:
                    # It's text like "n/a"
                    clean_price = "0"
                    price_note = f"Price: {raw_price}"
            except:
                clean_price = "0"
                price_note = f"Price: {raw_price}"

        # 5. MOQ
        raw_moq = clean_text(get_val(idx_moq))
        clean_moq = "1"
        if raw_moq:
            clean_moq = raw_moq

        # 6. MODEL
        final_model = clean_text(get_val(idx_model))

        # 7. NOTES
        raw_note = clean_text(get_val(idx_notes))
        notes_parts = []
        if price_note: notes_parts.append(price_note)
        if raw_note and raw_note.lower() != "nan": notes_parts.append(raw_note)
        
        final_notes = " | ".join(notes_parts)

        products.append({
            "Supplier": "ZOTAC",
            "Category": "ZOTAC VGA", # Hardcoded Sheet Name
            "Brand": "Zotac",
            "Model": final_model,
            "Name": final_name,
            "Price": clean_price,
            "Currency": "USD", 
            "Stock": "0", # Default (Not specified)
            "MOQ": clean_moq,
            "Notes": final_notes
        })

    # Output
    df_out = pd.DataFrame(products)
    
    if df_out.empty:
        print("Warning: No products found for ZOTAC.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from ZOTAC.")