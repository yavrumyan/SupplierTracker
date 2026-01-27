"""
conversion_logic_lans.py
Standardizes LANS's price-lists.
Logic:
- IGNORES Hidden Rows.
- Fixes line breaks (\n).
- Default Stock = 1.
- Currency = AMD.
- Header = Row 5.
- Column A = Name.
- Column B = Price (AMD).
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """
    Converts Excel column letter to 0-based index.
    A -> 0, B -> 1, etc.
    """
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def get_sheet_config(sheet_name):
    # LANS has only one sheet (or treats all sheets the same)
    return {
        "header_row": 5, 
        "indices": {
            "brand": [], # No Brand column specified
            "name": [col_index("A")],
            "price": [col_index("B")],
            "notes": []
        }
    }

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    """
    Standardize LANS price list.
    """
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'with_price': 0,
        'without_price': 0,
        'skipped_hidden': 0,
        'skipped_no_name': 0
    }

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        
        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    stats['skipped_hidden'] += 1
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. BUILD NAME FIRST
            name_cols = indices['name']
            name_parts = [clean_text(get_val(i)) for i in name_cols if get_val(i)]
            final_name = " ".join(name_parts)
            
            # Skip rows without name
            if not final_name:
                continue
            
            # Skip generic headers if they appear in data rows
            if any(keyword in final_name.upper() for keyword in ['MODEL NAME', 'PRICE', 'ЦЕНА', 'NAME']):
                continue

            stats['skipped_no_name'] += 1 if not final_name else 0

            # 3. EXTRACT PRICE
            price_cols = indices['price']
            raw_price = ""
            for p_idx in price_cols:
                p_val = get_val(p_idx)
                if p_val and p_val != "0":
                    raw_price = p_val
                    break 
            
            # Process price
            clean_price = "0"
            has_price = False
            
            if raw_price:
                try:
                    numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                    if numeric_str:
                        clean_price = str(int(float(numeric_str)))
                        has_price = True
                except:
                    if "call" in raw_price.lower():
                        clean_price = "CALL"
                        has_price = True
                    else:
                        clean_price = "0"
            
            if has_price and clean_price != "0":
                stats['with_price'] += 1
            else:
                stats['without_price'] += 1
            
            # Skip if no price and include_no_price is False
            if not include_no_price and (not has_price or clean_price == "0"):
                continue

            # 4. BUILD NOTES
            final_notes = ""
            if not has_price or clean_price == "0":
                final_notes = "Price Not Specified"

            products.append({
                "Supplier": "LANS",
                "Category": clean_text(sheet_name), 
                "Brand": "", 
                "Model": "",
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD", 
                "Stock": "1",
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for LANS.")
        return

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    
    print(f"Success! Converted {len(df_out)} items from LANS.")
    print(f"  - Products with price: {stats['with_price']}")
    print(f"  - Products without price: {stats['without_price']}")
    if stats['skipped_hidden'] > 0:
        print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")