"""
conversion_logic_kosatec.py
Standardizes KOSATEC's price-lists (Excel Version).
Logic:
- Input: .xlsx file.
- Fetches real-time EUR->USD rate from Frankfurter API.
- Converts Price (EUR) to USD.
- Rounds Price to 2 decimal places.
- Fixes line breaks.
- IGNORES Hidden Rows.
- Mapping:
    - Col B (Index 1) = Model
    - Col C (Index 2) = Name
    - Col D (Index 3) = Brand
    - Col G (Index 6) = Price (EUR -> USD)
    - Col J (Index 9) = Stock
    - "eta" column (Dynamic find) = Notes
"""

import pandas as pd
import openpyxl
import requests
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def get_eur_to_usd_rate():
    """
    Fetches the current EUR to USD exchange rate from Frankfurter API.
    Returns a float (e.g., 1.08).
    Returns a default fallback if API fails.
    """
    api_url = "https://api.frankfurter.dev/v1/latest?from=EUR&to=USD"
    fallback_rate = 1.18 # Approximate rate for 2026 fallback
    
    try:
        response = requests.get(api_url, timeout=5)
        response.raise_for_status()
        data = response.json()
        rate = data['rates']['USD']
        print(f"Fetched Exchange Rate (EUR->USD): {rate}")
        return float(rate)
    except Exception as e:
        print(f"Warning: Could not fetch exchange rate ({e}). Using fallback: {fallback_rate}")
        return fallback_rate

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def find_eta_column_index(ws, header_row):
    """
    Scans the header row to find the column named 'eta' (case-insensitive).
    Returns 0-based index or None.
    """
    for cell in ws[header_row]:
        if cell.value and "eta" in str(cell.value).lower():
            return cell.col_idx - 1 # openpyxl is 1-based, we return 0-based
    return None

def standardize(input_path: str, output_path: str):
    print("Starting KOSATEC conversion (Excel Version)...")
    
    # 1. Fetch Exchange Rate
    exchange_rate = get_eur_to_usd_rate()
    
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    
    # Process active sheet (usually just one for Kosatec)
    ws = wb.active
    
    # Configuration
    header_row = 1
    start_row = header_row + 1
    
    # Indices (0-based)
    # B=1, C=2, D=3, G=6, J=9
    idx_model = 1
    idx_name = 2
    idx_brand = 3
    idx_price = 6
    idx_stock = 9
    
    # Find ETA column dynamically
    idx_eta = find_eta_column_index(ws, header_row)
    
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        # 1. SKIP HIDDEN ROWS
        try:
            if ws.row_dimensions[row[0].row].hidden:
                continue
        except:
            pass

        cells = [cell.value for cell in row]
        
        def get_val(idx):
            if idx is None: return ""
            if 0 <= idx < len(cells):
                val = cells[idx]
                return str(val).strip() if val is not None else ""
            return ""

        # 2. NAME
        final_name = clean_text(get_val(idx_name))
        if not final_name:
            continue

        # 3. PRICE (EUR -> USD)
        raw_price = get_val(idx_price)
        clean_price = "0"
        
        try:
            # Clean format: "1.234,56" or "123,45"
            clean_str = raw_price.replace('€', '').strip()
            
            # If comma is decimal separator (European)
            if ',' in clean_str:
                 # Remove thousands dot if present (e.g. 1.200,00)
                if '.' in clean_str and clean_str.find('.') < clean_str.find(','):
                    clean_str = clean_str.replace('.', '')
                
                clean_str = clean_str.replace(',', '.')
            
            # Remove any remaining non-numeric chars except dot
            clean_str = "".join(c for c in clean_str if c.isdigit() or c == '.')
            
            if clean_str:
                val_eur = float(clean_str)
                val_usd = val_eur * exchange_rate
                
                val_usd = round(val_usd, 2)
                
                if val_usd.is_integer():
                    clean_price = str(int(val_usd))
                else:
                    clean_price = str(val_usd)
        except:
            clean_price = "0"

        # 4. STOCK
        raw_stock = get_val(idx_stock)
        clean_stock = "0"
        try:
            val = float(raw_stock)
            clean_stock = str(int(val))
        except:
            clean_stock = "0"

        # 5. OTHER FIELDS
        final_model = clean_text(get_val(idx_model))
        final_brand = clean_text(get_val(idx_brand))
        
        # Notes from ETA column
        final_notes = ""
        if idx_eta is not None:
            raw_eta = get_val(idx_eta)
            if raw_eta and raw_eta.lower() != "nan":
                final_notes = clean_text(raw_eta)

        products.append({
            "Supplier": "KOSATEC",
            "Category": "General", 
            "Brand": final_brand,
            "Model": final_model,
            "Name": final_name,
            "Price": clean_price,
            "Currency": "USD", 
            "Stock": clean_stock,
            "MOQ": "NO",
            "Notes": final_notes
        })

    # Output
    df_out = pd.DataFrame(products)
    
    if df_out.empty:
        print("Warning: No products found for KOSATEC.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from KOSATEC using Rate {exchange_rate}.")