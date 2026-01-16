"""
conversion_logic_elcore.py
Standardizes ELCORE's price-lists.
Filter-Proof: If filters hide the data, it automatically falls back to reading everything.
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text:
        return ""
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    s_name = sheet_name.strip()
    
    # Ignored Sheets
    if s_name in ["Main", "Contacts", "Сервисные центры", "Полезные ссылки", "Solar Systems"]:
        return None

    # Configuration Map
    if s_name == "Ноутбуки":
        return {"header_row": 1, "indices": {"model": 0, "brand": 1, "name": list(range(2, 13)), "price": 14, "notes": [15], "stock": 16}}
    elif s_name == "Моноблоки и Копьютеры":
        return {"header_row": 3, "indices": {"model": 0, "brand": 1, "name": list(range(2, 15)), "price": 16, "notes": [17], "stock": 18}}
    elif s_name == "Печатная техника":
        return {"header_row": 1, "indices": {"model": 0, "brand": None, "name": list(range(1, 10)), "price": 11, "notes": [13], "stock": 12}}
    elif s_name == "Мониторы":
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1, 2, 3, 4, 5, 6, 7, 13, 14, 17, 18], "price": 20, "notes": [21], "stock": 22}}
    elif s_name == "Планшеты":
        return {"header_row": 4, "indices": {"model": 0, "brand": 1, "name": list(range(2, 9)), "price": 10, "notes": [11], "stock": 12}}
    elif s_name == "Интерактивные дисплеи":
        return {"header_row": 3, "indices": {"model": 1, "brand": None, "name": [0, 2], "price": 5, "notes": [6], "stock": 3}}
    elif "UPS" in s_name:
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1, 2, 3, 6, 7, 8], "price": 12, "notes": [14, 15], "stock": 13}}
    elif s_name == "Xiaomi_ECO":
        return {"header_row": 3, "indices": {"model": 2, "brand": 0, "name": [1, 3], "price": 5, "notes": [6], "stock": 7}}
    elif s_name == "Смартфоны":
        return {"header_row": 3, "indices": {"model": 0, "brand": 1, "name": list(range(2, 7)), "price": 8, "notes": [9], "stock": 10}}
    elif s_name == "OEM":
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1], "price": 3, "notes": None, "stock": 4}}
    return None

def extract_row_data(row_cells, config, sheet_name):
    """Core logic to extract data from a row (list of values)."""
    indices = config['indices']
    
    # Helper to get value safely
    def get_val(idx):
        if idx is None: return ""
        if 0 <= idx < len(row_cells):
            val = row_cells[idx]
            return str(val).strip() if val is not None and str(val).lower() != 'nan' else ""
        return ""

    # 1. Price
    raw_price = get_val(indices['price'])
    if not raw_price: return None

    clean_price = "0"
    try:
        numeric_str = "".join(c for c in raw_price if c.isdigit())
        if numeric_str: clean_price = str(int(numeric_str))
    except: clean_price = "0"

    if clean_price == "0" and "call" not in raw_price.lower(): pass

    # 2. Name
    name_idx = indices['name']
    if isinstance(name_idx, list):
        name_parts = [clean_text(get_val(i)) for i in name_idx if get_val(i)]
        final_name = " ".join(name_parts)
    else:
        final_name = clean_text(get_val(name_idx))
    
    if not final_name: return None

    # 3. Notes
    notes_idx = indices.get('notes')
    final_notes = ""
    if notes_idx:
        if isinstance(notes_idx, list):
            final_notes = ", ".join([clean_text(get_val(i)) for i in notes_idx if get_val(i)])
        else:
            final_notes = clean_text(get_val(notes_idx))

    # 4. Stock
    raw_stock = get_val(indices.get('stock'))
    clean_stock = "0"
    if raw_stock:
        try: 
            val = float(raw_stock)
            clean_stock = str(int(val))
        except: 
            if raw_stock != "0": clean_stock = "1"

    return {
        "Supplier": "ELCORE",
        "Category": clean_text(sheet_name),
        "Brand": clean_text(get_val(indices.get('brand'))),
        "Model": clean_text(get_val(indices['model'])),
        "Name": final_name,
        "Price": clean_price,
        "Currency": "AMD",
        "Stock": clean_stock,
        "MOQ": "NO",
        "Notes": final_notes
    }

def standardize(input_path: str, output_path: str):
    all_products = []
    
    # --- METHOD 1: SMART (OpenPyXL) ---
    # Attempts to skip hidden rows (Standard behavior).
    print("Attempting Method 1 (Smart Hidden-Row Detection)...")
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        for sheet_name in wb.sheetnames:
            config = get_sheet_config(sheet_name)
            if not config: continue
            
            ws = wb[sheet_name]
            start_row = config['header_row'] + 1
            
            for row in ws.iter_rows(min_row=start_row, values_only=False):
                if not row: continue
                # Check Hidden
                try:
                    if ws.row_dimensions[row[0].row].hidden: continue
                except: pass

                cells = [c.value for c in row]
                item = extract_row_data(cells, config, sheet_name)
                if item: all_products.append(item)
                
    except Exception as e:
        print(f"Method 1 failed: {e}")

    # --- METHOD 2: FALLBACK (Brute Force) ---
    # If Method 1 was blocked by filters (result count is 0), 
    # we switch to Pandas which IGNORES filters and reads everything.
    if len(all_products) == 0:
        print("Method 1 yielded 0 results (likely due to Excel Filters). Switching to Method 2 (Reading All Data)...")
        try:
            xls_dict = pd.read_excel(input_path, sheet_name=None, header=None)
            for sheet_name, df in xls_dict.items():
                config = get_sheet_config(sheet_name)
                if not config: continue
                
                header_idx = config['header_row'] - 1
                if len(df) <= header_idx: continue
                
                # Iterate over raw pandas rows
                for _, row in df.iloc[header_idx+1:].iterrows():
                    cells = row.tolist()
                    item = extract_row_data(cells, config, sheet_name)
                    if item: all_products.append(item)
                    
        except Exception as e:
            print(f"Method 2 failed: {e}")

    # Output
    df_out = pd.DataFrame(all_products)
    
    if df_out.empty:
        # Create dummy structure to avoid system crash if truly empty
        print("CRITICAL: No products found. Creating empty template.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # Force string type for JSON compatibility
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Done. Total products: {len(df_out)}")