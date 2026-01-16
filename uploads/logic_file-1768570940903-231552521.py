"""
conversion_logic_armit.py
Standardizes ARMIT's price-lists (ARMIT.xlsx).
Fixes:
- Syntax Error "unterminated string literal" resolved.
- Removes newlines/hidden formatting safely.
- Ignores Hidden rows.
"""

import pandas as pd
import openpyxl

def clean_text(text):
    """
    Safe helper to remove line breaks and extra spaces.
    Uses split() to handle all whitespace (newlines, tabs) automatically.
    """
    if not text:
        return ""
    # split() splits by ANY whitespace (space, tab, enter) and removes empty parts
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns the configuration for a given sheet name.
    """
    s_name = sheet_name.strip().upper()

    if s_name == "MAIN":
        return None 

    # 1. Sheets: NETWORK, PRINTERS, SERVERS
    if s_name in ["NETWORK", "PRINTERS", "SERVERS"]:
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 2. Sheets: LAPTOPS, PC | AIO
    elif s_name in ["COMMERCIAL LAPTOPS", "CONSUMER LAPTOPS", "PC | AIO"]:
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None, "name": [1], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 3. Sheet: MONITORS
    elif s_name == "MONITORS":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2, 3, 4, 5, 6, 7, 8], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 4. Sheet: FSP UPS
    elif s_name == "FSP UPS":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2, 3, 4, 5, 6, 7], 
                "price": 8, "stock": 9, "notes": [10]
            }
        }

    # 5. Sheet: PROJECTORS
    elif "PROJECTORS" in s_name:
        return {
            "header_row": 13,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 6. Sheet: SYNOLOGY
    elif s_name == "SYNOLOGY":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2, 3], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 7. Sheet: ACCESSORIES
    elif s_name == "ACCESSORIES":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": 1, "name": [2, 3], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 8. Sheet: PC COMPONENTS
    elif s_name == "PC COMPONENTS":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": 1, "name": [2, 3, 4, 5, 6, 7, 8], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 9. Sheet: QSAN
    elif s_name == "QSAN":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2, 3, 4, 5, 6, 7, 8], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    # 10. Sheet: INTERACTIVE DISPLAYS
    elif s_name == "INTERACTIVE DISPLAYS":
        return {
            "header_row": 5,
            "indices": {
                "model": 0, "brand": None, "name": [1, 2], 
                "price": 9, "stock": 10, "notes": [11]
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SAFETY: Skip empty rows
            if not row:
                continue

            # 2. IGNORE HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 3. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass 

            # 4. BUILD NAME (Clean Newlines)
            name_parts = [clean_text(get_val(i)) for i in indices['name'] if get_val(i)]
            final_name = " ".join(name_parts)
            
            if not final_name:
                continue

            # 5. BUILD NOTES & STOCK
            notes_raw = [clean_text(get_val(i)) for i in indices['notes'] if get_val(i)]
            final_notes = ", ".join(notes_raw)
            
            raw_stock = get_val(indices['stock'])
            clean_stock = "1"
            if raw_stock:
                try:
                    clean_stock = str(int(float(raw_stock)))
                except:
                    clean_stock = "1"
            
            products.append({
                "Supplier": "ARMIT",
                "Category": clean_text(sheet_name),
                "Brand": clean_text(get_val(indices['brand'])),
                "Model": clean_text(get_val(indices['model'])),
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD",
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for ARMIT.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from ARMIT.")