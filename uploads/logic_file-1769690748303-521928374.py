"""
conversion_logic_phonix.py
Standardizes PHONIX's price-lists.
Logic:
- Input: .xlsx file.
- IGNORES Sheets: "TERMS AND CONDITIONS", "Just Launched", "Today's Deals", "Price Drop", "Gaming".
- PROCESSES all other sheets (In Stock, Laptops, etc.).
- Sheet Name = Category.
- IGNORES Hidden Rows.
- Fixes line breaks.
- ROUNDS PRICES to 2 decimal places.
- Mapping:
    - Col A = Brand
    - Col B = Model
    - Col C + D = Name
    - Col E = Price (USD)
    - Col H = Stock
    - Notes = "MOQ " + Col F + " ETA " + Col G + " Conditions " + Col I + Col J
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def standardize(input_path: str, output_path: str):
    print("Starting PHONIX conversion...")
    
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    
    # Sheets to explicitly ignore (lowercase for comparison)
    IGNORE_SHEETS = [
        "terms and conditions",
        "just launched",
        "today's deals", 
        "price drop", 
        "gaming"
    ]

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in IGNORE_SHEETS:
            continue

        ws = wb[sheet_name]
        
        # Configuration
        # Header = Row 1, Data starts Row 2
        start_row = 2
        
        # Column Indices (0-based)
        # A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9
        idx_brand = 0
        idx_model = 1
        idx_name_1 = 2
        idx_name_2 = 3
        idx_price = 4
        idx_moq = 5
        idx_eta = 6
        idx_stock = 7
        idx_condition = 8
        idx_note_extra = 9

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. NAME (Col C + Col D)
            part1 = clean_text(get_val(idx_name_1))
            part2 = clean_text(get_val(idx_name_2))
            final_name = f"{part1} {part2}".strip()
            
            if not final_name:
                continue

            # 3. PRICE (USD)
            raw_price = get_val(idx_price)
            clean_price = "0"
            
            try:
                # Remove '$' and ',' then convert
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    val = float(numeric_str)
                    val = round(val, 2) # Round to 2 decimals
                    
                    if val.is_integer():
                        clean_price = str(int(val))
                    else:
                        clean_price = str(val)
            except:
                if "call" in raw_price.lower():
                    clean_price = "CALL"
                else:
                    clean_price = "0"

            # 4. STOCK (Col H)
            raw_stock = get_val(idx_stock)
            clean_stock = "0"
            try:
                numeric_stock = "".join(c for c in raw_stock if c.isdigit())
                if numeric_stock:
                    clean_stock = str(int(numeric_stock))
            except:
                clean_stock = "0"

            # 5. BRAND & MODEL
            final_brand = clean_text(get_val(idx_brand))
            final_model = clean_text(get_val(idx_model))

            # 6. NOTES (Complex Merge)
            # Structure: MOQ {F} | ETA {G} | Conditions {I} | {J}
            notes_parts = []
            
            val_moq = clean_text(get_val(idx_moq))
            if val_moq: notes_parts.append(f"MOQ {val_moq}")
            
            val_eta = clean_text(get_val(idx_eta))
            if val_eta: notes_parts.append(f"ETA {val_eta}")
            
            val_cond = clean_text(get_val(idx_condition))
            if val_cond: notes_parts.append(f"Conditions {val_cond}")
            
            val_extra = clean_text(get_val(idx_note_extra))
            if val_extra: notes_parts.append(val_extra)
            
            final_notes = " | ".join(notes_parts)

            products.append({
                "Supplier": "PHONIX",
                "Category": clean_text(sheet_name), # Sheet Name = Category
                "Brand": final_brand,
                "Model": final_model,
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", 
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Output
    df_out = pd.DataFrame(products)
    
    if df_out.empty:
        print("Warning: No products found for PHONIX.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from PHONIX.")