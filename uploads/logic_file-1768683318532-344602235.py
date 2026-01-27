"""
conversion_logic_dg.py
Standardizes DG's price-lists.
Logic:
- IGNORES Hidden Rows.
- Fixes line breaks.
- ROUNDS PRICES to 2 decimal places.
- Extracts Brand/Model from Product Name.
- Header = Row 2.
- Columns: A=Name, B=Stock, C=Price(USD).
"""

import pandas as pd
import openpyxl
import warnings
import re

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def extract_brand_from_name(product_name):
    """
    Extract brand name from product description.
    """
    if not product_name: return "Unknown"
    
    # Common brands to look for
    brands = [
        'Lenovo', 'Dell', 'Asus', 'ASUS', 'HP', 'Acer', 'MSI', 'Gigabyte',
        'DeepCool', 'Intel', 'AMD', 'Nvidia', 'Samsung', 'LG', 'BenQ',
        'ViewSonic', 'AOC', 'Philips', 'Canon', 'Epson', 'Brother',
        'APC', 'CyberPower', 'Eaton', 'Schneider', 'Logitech', 'Razer',
        'Corsair', 'Kingston', 'WD', 'Seagate', 'Transcend', 'SanDisk',
        'TP-Link', 'D-Link', 'Netgear', 'Cisco', 'Ubiquiti', 'MikroTik'
    ]
    
    product_upper = product_name.upper()
    for brand in brands:
        if brand.upper() in product_upper:
            return brand
    
    return "Unknown"

def extract_model_from_name(product_name):
    """
    Try to extract model number from product description.
    """
    if not product_name: return ""
    
    patterns = [
        r'\b([A-Z0-9]+-[A-Z0-9-]+)\b',  # e.g., USB-N10, A620M-K
        r'\b([A-Z]+\d+[A-Z]*)\b',        # e.g., V130, 7400
        r'\b(\d{4,}[A-Z]*)\b',            # e.g., 7400, 5400
    ]
    
    for pattern in patterns:
        match = re.search(pattern, product_name)
        if match:
            return match.group(1)
    return ""

def get_sheet_config(sheet_name):
    # DG structure appears consistent across sheets
    # Header = Row 2 (Index 1) -> Data starts Row 3
    # Col A (0) = Name
    # Col B (1) = Stock
    # Col C (2) = Price
    return {
        "header_row": 2,  
        "indices": {
            "name": [0],
            "stock": 1,
            "price": 2
        }
    }

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'processed': 0,
        'skipped_hidden': 0
    }

    for sheet_name in wb.sheetnames:
        # Skip generic generic sheet names if they don't contain data
        # But DG usually puts categories in sheet names, so we process all
        config = get_sheet_config(sheet_name)
        
        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    stats['skipped_hidden'] += 1
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. NAME (Required)
            name_parts = [clean_text(get_val(i)) for i in indices['name'] if get_val(i)]
            final_name = " ".join(name_parts)
            
            # Skip empty rows or Date/Header rows that slipped through
            if not final_name or "Անվանում" in final_name or "USD" in final_name:
                continue

            # 3. PRICE (Round to 2 decimals)
            raw_price = get_val(indices['price'])
            clean_price = "0"
            
            if raw_price:
                try:
                    numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                    if numeric_str:
                        val = float(numeric_str)
                        val = round(val, 2)
                        
                        if val.is_integer():
                            clean_price = str(int(val))
                        else:
                            clean_price = str(val)
                except:
                    if "call" in raw_price.lower():
                        clean_price = "CALL"
                    else:
                        clean_price = "0"

            # 4. STOCK
            raw_stock = get_val(indices['stock'])
            clean_stock = "0"
            if raw_stock:
                try:
                    # Sanitize stock (remove <, >, etc)
                    numeric_stock = "".join(c for c in raw_stock if c.isdigit() or c == '.')
                    if numeric_stock:
                        clean_stock = str(int(float(numeric_stock)))
                except:
                    if raw_stock.lower() not in ["0", "-", "no", "absent"]:
                        clean_stock = "1"

            # 5. BRAND & MODEL (Extracted)
            final_brand = extract_brand_from_name(final_name)
            final_model = extract_model_from_name(final_name)

            # 6. NOTES
            final_notes = ""
            if clean_price == "0":
                final_notes = "Price Not Specified"

            products.append({
                "Supplier": "DG",
                "Category": clean_text(sheet_name), 
                "Brand": final_brand, 
                "Model": final_model,
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", 
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })
            stats['processed'] += 1

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for DG.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from DG.")
    print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")