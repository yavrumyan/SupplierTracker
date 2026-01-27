"""
conversion_logic_complanet.py
Standardizes COMPLANET's price-lists.
Logic:
- IGNORES Hidden Rows.
- IGNORES Sheets not in the Allowlist.
- Fixes line breaks (\n).
- Default Stock = 1.
- Currency = USD (inferred from values like 3, 920, 120).
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """
    Converts Excel column letter to 0-based index.
    A -> 0, B -> 1, Z -> 25, AA -> 26, etc.
    """
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def get_sheet_config(sheet_name):
    s_name = sheet_name.strip()
    
    # Helper to generate range of indices from start col to end col (inclusive)
    def col_range(start_col, end_col):
        start_idx = col_index(start_col)
        end_idx = col_index(end_col)
        return list(range(start_idx, end_idx + 1))

    # Helper for specific list of cols
    def cols(col_list):
        return [col_index(c) for c in col_list]

    # --- CONFIGURATION ---
    
    if s_name == "NOTEBOOKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),
                "notes": cols(["AB", "AC", "AF", "AG"])
            }
        }
        
    elif s_name == "MONITORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "ALL IN ONE PC":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "RAM":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "CPU":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),
                "notes": cols(["AB", "AC", "AF", "AG", "AH"])
            }
        }

    elif s_name == "HDD, SSD":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "CASE":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "MOTHER BOARDS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD", "AE"])
            }
        }

    elif s_name == "COOLER  FAN": # Note double space in sheet name as per user
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "POWER SYPLY":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "VGA":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "UPS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "PRINTERS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "DVD-RW":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"]) # T overlaps with Name, per user instruction
            }
        }

    elif s_name == "PROJECTORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NETWORKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"])
            }
        }

    elif s_name == "USB FLASH":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "N"),
                "price": cols(["Q", "R"]),
                "notes": cols(["O", "P", "S", "T", "U"])
            }
        }

    elif s_name == "PC COMPOMEMTS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NOTEBOOK BAGS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "R"),
                "price": cols(["U", "V"]),
                "notes": cols(["S", "T", "W", "X", "Y"])
            }
        }

    elif s_name == "SECURITY CAMERAS":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "Home appliances":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "L"),
                "price": cols(["O", "P"]),
                "notes": cols(["M", "N", "Q", "R", "S"])
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. EXTRACT PRICE
            # Check primary price col, then secondary
            price_cols = indices['price']
            raw_price = ""
            for p_idx in price_cols:
                p_val = get_val(p_idx)
                if p_val and p_val != "0":
                    raw_price = p_val
                    break # Take first valid price
            
            if not raw_price:
                continue

            clean_price = "0"
            try:
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    clean_price = str(int(float(numeric_str)))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass

            # 3. BUILD BRAND
            brand_cols = indices['brand']
            brand_parts = [clean_text(get_val(i)) for i in brand_cols if get_val(i)]
            final_brand = " ".join(brand_parts)

            # 4. BUILD NAME
            name_cols = indices['name']
            name_parts = [clean_text(get_val(i)) for i in name_cols if get_val(i)]
            final_name = " ".join(name_parts)
            
            if not final_name:
                continue

            # 5. BUILD NOTES
            notes_cols = indices.get('notes', [])
            notes_parts = [clean_text(get_val(i)) for i in notes_cols if get_val(i)]
            final_notes = ", ".join(notes_parts)

            products.append({
                "Supplier": "COMPLANET",
                "Category": clean_text(sheet_name),
                "Brand": final_brand, 
                "Model": "",
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", # Inferred USD
                "Stock": "1",
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for COMPLANET.")
        return

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from COMPLANET.")