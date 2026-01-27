"""
conversion_logic_brand.py
Standardizes BRAND's price-lists.
Logic:
- IGNORES Hidden Rows.
- IGNORES Sheets not in the Allowlist.
- Fixes line breaks (\n).
- Default Stock = 1.
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    """
    Removes line breaks, tabs, and extra spaces.
    """
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns configuration for a given sheet name.
    header_row: 1-based Excel row number.
    indices: 0-based column indices (Col A=0, B=1, etc).
    """
    s_name = sheet_name.strip()

    # --- Dell ---
    # Header 2. Name(1+3), Price(4)
    if s_name == "Dell":
        return {
            "header_row": 2,
            "indices": {
                "model": None, "brand": None,
                "name": [0, 2], # Col 1 is index 0
                "price": 3, "notes": None
            }
        }

    # --- LENOVO ---
    # Header 2. Name(1+3), Price(4)
    elif s_name == "LENOVO":
        return {
            "header_row": 2,
            "indices": {
                "model": None, "brand": None,
                "name": [0, 2], 
                "price": 3, "notes": None
            }
        }

    # --- Whiteboard ---
    # Header 1. Name(1+5), Price(6)
    elif s_name == "Whiteboard":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0, 4], 
                "price": 5, "notes": None
            }
        }

    # --- Deepcool ---
    # Header 1. Name(1), Price(3)
    elif s_name == "Deepcool":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0], 
                "price": 2, "notes": None
            }
        }

    # --- Xilence ---
    # Header 1. Name(1), Price(3)
    elif s_name == "Xilence":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0], 
                "price": 2, "notes": None
            }
        }

    # --- Genius ---
    # Header 1. Name(1), Price(2)
    elif s_name == "Genius":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0], 
                "price": 1, "notes": None
            }
        }

    # --- Tenda ---
    # Header 1. Name(2+3), Price(4)
    elif s_name == "Tenda":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [1, 2], 
                "price": 3, "notes": None
            }
        }

    # --- Gembird ---
    # Header 1. Name(1), Price(3)
    elif s_name == "Gembird":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0], 
                "price": 2, "notes": None
            }
        }

    # --- Gembird UPS ---
    # Header 1. Name(1+2), Price(4)
    elif s_name == "Gembird UPS":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0, 1], 
                "price": 3, "notes": None
            }
        }

    # --- APC ---
    # Header 1. Model(1), Name(1+2), Price(4)
    elif s_name == "APC":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": None,
                "name": [0, 1], 
                "price": 3, "notes": None
            }
        }

    # --- VSOL ---
    # Header 1. Model(1), Name(1+3), Price(4)
    elif s_name == "VSOL":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": None,
                "name": [0, 2], 
                "price": 3, "notes": None
            }
        }

    # --- InFocus ---
    # Header 1. Model(1), Name(1+7+8+9+10+11), Price(12)
    elif s_name == "InFocus":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": None,
                "name": [0, 6, 7, 8, 9, 10], 
                "price": 11, "notes": None
            }
        }

    # --- Network ---
    # Header 1. Name(1+4), Price(5)
    elif s_name == "Network":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0, 3], 
                "price": 4, "notes": None
            }
        }

    # --- Server Cabinets ---
    # Header 2. Model(1), Name(1+4), Price(5)
    elif s_name == "Server Cabinets":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [0, 3], 
                "price": 4, "notes": None
            }
        }

    # --- Other ---
    # Header 1. Name(1), Price(2)
    elif s_name == "Other":
        return {
            "header_row": 1,
            "indices": {
                "model": None, "brand": None,
                "name": [0], 
                "price": 1, "notes": None
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            # Skip ignored sheets (Title, etc.)
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        # Iterate Rows
        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                # Keep only digits
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass 

            # 3. BUILD NAME
            name_idx = indices['name']
            if isinstance(name_idx, list):
                name_parts = [clean_text(get_val(i)) for i in name_idx if get_val(i)]
                final_name = " ".join(name_parts)
            else:
                final_name = clean_text(get_val(name_idx))
            
            if not final_name:
                continue

            # 4. BUILD NOTES
            notes_idx = indices.get('notes')
            final_notes = ""
            if notes_idx:
                if isinstance(notes_idx, list):
                    final_notes = ", ".join([clean_text(get_val(i)) for i in notes_idx if get_val(i)])
                else:
                    final_notes = clean_text(get_val(notes_idx))

            products.append({
                "Supplier": "BRAND",
                "Category": clean_text(sheet_name),
                "Brand": "", 
                "Model": clean_text(get_val(indices.get('model'))),
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD",
                "Stock": "1", # Default to 1
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for BRAND.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from BRAND.")