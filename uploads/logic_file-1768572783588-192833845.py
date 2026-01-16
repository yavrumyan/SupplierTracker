"""
conversion_logic_it_plaza.py
Standardizes IT-PLAZA's price-lists (IT-PLAZA.xlsx).

Logic:
- Sheet Name = Category.
- IGNORES hidden rows.
- Fixes line breaks (\n) in cells.
- Varying Header Rows (Row 2, 4, 5, 6, 7).
- Varying Column Mappings.
"""

import pandas as pd
import openpyxl

def clean_text(text):
    """
    Safe helper to remove line breaks and extra spaces.
    """
    if not text:
        return ""
    # split() handles all whitespace (newlines, tabs) automatically
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns the configuration for a given sheet name.
    header_row is the Excel row number (1-based).
    indices are 0-based (Column A=0, B=1, C=2...).
    """
    s_name = sheet_name.strip()

    # 1. Sheet: "NB, AIO,NB Cases"
    # Header 7, Name Col 2 (1), Price Col 3 (2), Notes Col 4 (3)
    if s_name == "NB, AIO,NB Cases":
        return {
            "header_row": 7,
            "indices": {"name": 1, "price": 2, "notes": 3}
        }

    # 2. Sheet: "PC Components"
    # Header 7, Name Col 2 (1), Price Col 4 (3), Notes Col 5 (4)
    elif s_name == "PC Components":
        return {
            "header_row": 7,
            "indices": {"name": 1, "price": 3, "notes": 4}
        }

    # 3. Sheet: "PC Accessories"
    # Header 4, Name Col 2 (1), Price Col 4 (3), Notes Col 5 (4)
    elif s_name == "PC Accessories":
        return {
            "header_row": 4,
            "indices": {"name": 1, "price": 3, "notes": 4}
        }

    # 4. Sheet: "Network equipment"
    # Header 7, Name Col 2 (1), Price Col 3 (2)
    elif s_name == "Network equipment":
        return {
            "header_row": 7,
            "indices": {"name": 1, "price": 2, "notes": None}
        }

    # 5. Sheet: "TV & Monitor Holders, Boxes"
    # Header 5, Name Col 2 (1), Price Col 4 (3)
    elif s_name == "TV & Monitor Holders, Boxes":
        return {
            "header_row": 5,
            "indices": {"name": 1, "price": 3, "notes": None}
        }

    # 6. Sheet: "Printers & Projectors " (Note space handling)
    # Header 7, Name Col 2 (1), Price Col 3 (2), Notes Col 4 (3)
    elif "Printers & Projectors" in s_name:
        return {
            "header_row": 7,
            "indices": {"name": 1, "price": 2, "notes": 3}
        }

    # 7. Sheet: "Adapters & Cables"
    # Header 6, Name Col 2 (1), Price Col 3 (2)
    elif s_name == "Adapters & Cables":
        return {
            "header_row": 6,
            "indices": {"name": 1, "price": 2, "notes": None}
        }

    # 8. Sheet: "Furniture"
    # Header 2, Name Col 2 (1), Price Col 3 (2)
    elif s_name == "Furniture":
        return {
            "header_row": 2,
            "indices": {"name": 1, "price": 2, "notes": None}
        }

    # 9. Sheet: "battery"
    # Header 7, Name Col 2 (1), Price Col 3 (2)
    elif s_name == "battery":
        return {
            "header_row": 7,
            "indices": {"name": 1, "price": 2, "notes": None}
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load Workbook with openpyxl
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        # Iterate Rows
        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SAFETY: Skip empty rows
            if not row:
                continue

            # 2. IGNORE HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 3. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                # Remove non-digits (keep only numbers)
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass 

            # 4. BUILD NAME (Clean Newlines)
            final_name = clean_text(get_val(indices['name']))
            if not final_name:
                continue

            # 5. BUILD NOTES
            final_notes = ""
            if indices['notes'] is not None:
                final_notes = clean_text(get_val(indices['notes']))

            products.append({
                "Supplier": "IT-PLAZA",
                "Category": clean_text(sheet_name),
                "Brand": "", 
                "Model": "",
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD",
                "Stock": "1", # Default to 1
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for IT-PLAZA.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from IT-PLAZA.")