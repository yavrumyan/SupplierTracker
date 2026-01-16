"""
conversion_logic_meta.py
Standardizes META's price-lists (META.xlsx).

Logic:
- Sheet Name = Category.
- Header is Row 1 (Index 0).
- IGNORING HIDDEN ROWS (Critical).
- Different column mappings per sheet.
"""

import pandas as pd
import openpyxl

def get_sheet_config(sheet_name):
    """
    Returns the column mapping configuration for a given sheet name.
    Indices are 0-based (e.g., 13th column = index 12).
    """
    s_name = sheet_name.strip().lower()

    # Sheet: "Notebook"
    if s_name == "notebook":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 12, "notes": [13]
            },
            "currency": "USD"
        }

    # Sheet: "Monitor"
    elif s_name == "monitor":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 22, "notes": [23]
            },
            "currency": "USD"
        }

    # Sheet: "Honor"
    elif s_name == "honor":
        return {
            "indices": {
                "brand": 2, "model": 1, "name": 3, 
                "price": 4, "notes": [7]
            },
            "currency": "AMD"
        }

    # Sheet: "AIO"
    elif s_name == "aio":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 14, "notes": [15]
            },
            "currency": "USD"
        }

    # Sheet: "PC Components"
    elif s_name == "pc components":
        return {
            "indices": {
                "brand": 1, "model": 0, "name": [2, 3], # Combine 2+3
                "price": 4, "notes": [5]
            },
            "currency": "USD"
        }

    # Sheet: "Accessories"
    elif s_name == "accessories":
        return {
            "indices": {
                "brand": 2, "model": 0, "name": [3, 4], # Combine 3+4
                "price": 5, "notes": [6]
            },
            "currency": "USD"
        }

    # Sheet: "Xiaomi+Roborock"
    elif "xiaomi" in s_name and "roborock" in s_name:
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 3, "notes": [5]
            },
            "currency": "AMD"
        }

    # Sheet: "Dreame" (Handles "Dreame " with space)
    elif "dreame" in s_name:
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 3, "notes": [5, 6] # Combine 5+6
            },
            "currency": "AMD"
        }

    # Sheet: "Mova"
    elif s_name == "mova":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 3, "notes": [6, 7] # Combine 6+7
            },
            "currency": "AMD"
        }

    # Sheet: "HUT"
    elif s_name == "hut":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 3, "notes": [4]
            },
            "currency": "USD"
        }

    # Sheet: "Viaomi"
    elif s_name == "viaomi":
        return {
            "indices": {
                "brand": 0, "model": 1, "name": 2, 
                "price": 3, "notes": [4]
            },
            "currency": "USD"
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load Workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            # Skip sheets we don't know (or Summary sheets)
            continue

        ws = wb[sheet_name]
        
        # Iterate Rows
        # min_row=2 skips the Header (Row 1)
        for row in ws.iter_rows(min_row=2, values_only=False):
            # 1. CHECK HIDDEN STATUS
            # If the row is hidden in Excel, skip it entirely
            if ws.row_dimensions[row[0].row].hidden:
                continue

            # Extract cell values
            cells = [cell.value for cell in row]
            
            # Helper to safely get value by index
            def get_val(idx):
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            indices = config['indices']
            
            # 2. EXTRACT & CLEAN PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue # Skip if no price

            # Remove non-numeric chars (except digits)
            # This handles "150 USD" -> "150" or "Call" -> "0"
            clean_price = "0"
            try:
                # Filter string to keep only digits and dots
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    clean_price = str(int(float(numeric_str)))
            except:
                clean_price = "0"
            
            if clean_price == "0" and "call" not in raw_price.lower():
                 # If price became 0 but wasn't "Call", and wasn't empty, it might be junk data
                 pass

            # 3. BUILD NAME (Merge columns if needed)
            name_idx = indices['name']
            if isinstance(name_idx, list):
                # Combine multiple columns
                name_parts = [get_val(i) for i in name_idx if get_val(i)]
                final_name = " ".join(name_parts)
            else:
                final_name = get_val(name_idx)

            if not final_name: 
                continue

            # 4. BUILD NOTES (Merge columns if needed)
            notes_idx = indices['notes']
            note_parts = [get_val(i) for i in notes_idx if get_val(i)]
            final_notes = ", ".join(note_parts)

            products.append({
                "Supplier": "META",
                "Category": sheet_name.strip(), # Original Case Sheet Name
                "Brand": get_val(indices['brand']),
                "Model": get_val(indices['model']),
                "Name": final_name,
                "Price": clean_price,
                "Currency": config['currency'],
                "Stock": "1", # Default to 1
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for META.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from META.")