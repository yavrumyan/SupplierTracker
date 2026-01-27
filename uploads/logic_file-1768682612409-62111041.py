"""
conversion_logic_muk.py
Standardizes MUK's price-lists.
Logic:
- IGNORES Hidden Rows.
- Fixes line breaks.
- DYNAMIC HEADER DETECTION (Scans for 'Price'/'Model').
- ROUNDS PRICES to 2 decimal places (e.g. 27.62).
- Default Stock = 1 (if not specified).
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """Converts Excel column letter to 0-based index."""
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def find_header_row(ws):
    """
    Scans the first 20 rows to find the header.
    Looks for row containing 'Model' AND 'Price'.
    """
    for r in range(1, 21):
        row_str = ""
        for c in range(1, 15):
            val = ws.cell(row=r, column=c).value
            if val: row_str += str(val).upper() + " "
        
        if "MODEL" in row_str and ("PRICE" in row_str or "DEALER" in row_str or "SRP" in row_str):
            return r
    return None

def get_sheet_config(sheet_name, header_row):
    """
    Defines column mappings based on standard MUK layout.
    Adjust indices if your specific MUK file differs.
    """
    # Standard MUK Layout often:
    # Col A/B = Category/Brand
    # Col C/D = Model/Part Num
    # Col E/F = Name/Description
    # Col G/H = Price
    
    # We'll use a generic mapping relative to found header, 
    # but specific indices usually need to be hardcoded or detected.
    # Assuming standard MUK structure here:
    return {
        "indices": {
            "brand": 1,    # Col B (often Brand)
            "model": 2,    # Col C (often Part Number)
            "name": [3],   # Col D (Description)
            "price": 6,    # Col G (Dealer Price/Price)
            "stock": 5,    # Col F (Stock/Qty)
            "notes": [8]   # Col I (Comments/Notes)
        }
    }

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'processed': 0,
        'skipped_hidden': 0
    }

    for sheet_name in wb.sheetnames:
        # MUK usually has data in "Price" or "Sheet1", skip obvious junk
        if "sheet" not in sheet_name.lower() and "price" not in sheet_name.lower() and "list" not in sheet_name.lower():
             # Heuristic: if it looks like a cover page, skip. 
             # But safer to try processing if header is found.
             pass

        ws = wb[sheet_name]
        
        # 1. FIND HEADER
        header_row = find_header_row(ws)
        if not header_row:
            continue # Skip sheets without a valid header

        config = get_sheet_config(sheet_name, header_row)
        start_row = header_row + 1
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 2. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    stats['skipped_hidden'] += 1
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 3. NAME (Required)
            name_parts = [clean_text(get_val(i)) for i in indices['name'] if get_val(i)]
            final_name = " ".join(name_parts)
            if not final_name:
                continue
            
            # Skip repetition of headers
            if "MODEL" in final_name.upper() and "PRICE" in final_name.upper():
                continue

            # 4. PRICE (Rounded to 2 decimals)
            raw_price = get_val(indices['price'])
            clean_price = "0"
            
            if raw_price:
                try:
                    numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                    if numeric_str:
                        val = float(numeric_str)
                        val = round(val, 2) # Round to 2 decimals
                        
                        if val.is_integer():
                            clean_price = str(int(val))
                        else:
                            clean_price = str(val)
                except:
                    if "call" in raw_price.lower():
                        clean_price = "CALL"
                    else:
                        clean_price = "0"

            # 5. STOCK (Handle > symbols)
            raw_stock = get_val(indices['stock'])
            clean_stock = "0"
            if raw_stock:
                try:
                    val = float(raw_stock)
                    clean_stock = str(int(val))
                except:
                    # Handle ">10"
                    numeric_stock = "".join(c for c in raw_stock if c.isdigit() or c == '.')
                    if numeric_stock:
                        clean_stock = str(int(float(numeric_stock)))
                    else:
                        if raw_stock.lower() not in ["0", "-", "no"]:
                            clean_stock = "1"

            # 6. BRAND & MODEL
            final_brand = clean_text(get_val(indices['brand']))
            final_model = clean_text(get_val(indices['model']))
            
            # 7. NOTES
            notes_parts = [clean_text(get_val(i)) for i in indices['notes'] if get_val(i)]
            if clean_price == "0":
                notes_parts.append("Price Not Specified")
            final_notes = ", ".join(notes_parts)

            products.append({
                "Supplier": "MUK",
                "Category": clean_text(sheet_name), 
                "Brand": final_brand, 
                "Model": final_model,
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", # MUK usually USD, verify if needed
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })
            stats['processed'] += 1

    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for MUK.")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from MUK.")
    print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")