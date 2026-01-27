"""
conversion_logic_complanet.py
Standardizes COMPLANET's price-lists.
Logic:
- IGNORES Hidden Rows.
- IGNORES Sheets not in the Allowlist.
- Fixes line breaks (\n).
- Default Stock = 1.
- Currency = USD (Confirmed).
- INCLUDES products even if they have NO PRICE (sets price to "0" or "TBD")
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """
    Converts Excel column letter to 0-based index.
    A -> 0, B -> 1, Z -> 25, AA -> 26, etc.
    """
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def get_sheet_config(sheet_name):
    s_name = sheet_name.strip()
    
    # Helper to generate range of indices
    def col_range(start_col, end_col):
        start_idx = col_index(start_col)
        end_idx = col_index(end_col)
        return list(range(start_idx, end_idx + 1))

    # Helper for specific list of cols
    def cols(col_list):
        return [col_index(c) for c in col_list]

    # --- CONFIGURATION ---
    
    if s_name == "NOTEBOOKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),  # AD=Price with NDS (VAT included) - this is what we want
                "notes": cols(["AB", "AC", "AF", "AG"])  # AB=Price without VAT goes to notes
            }
        }
        
    elif s_name == "MONITORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "ALL IN ONE PC":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "RAM":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "CPU":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),
                "notes": cols(["AB", "AC", "AF", "AG", "AH"])
            }
        }

    elif s_name == "HDD, SSD":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "CASE":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "MOTHER BOARDS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD", "AE"])
            }
        }

    elif s_name == "COOLER  FAN":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "POWER SYPLY":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "VGA":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "UPS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "PRINTERS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "DVD-RW":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"])
            }
        }

    elif s_name == "PROJECTORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NETWORKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"])
            }
        }

    elif s_name == "USB FLASH":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "N"),
                "price": cols(["Q", "R"]),
                "notes": cols(["O", "P", "S", "T", "U"])
            }
        }

    elif s_name == "PC COMPOMEMTS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NOTEBOOK BAGS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "R"),
                "price": cols(["U", "V"]),
                "notes": cols(["S", "T", "W", "X", "Y"])
            }
        }

    elif s_name == "SECURITY CAMERAS":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "Home appliances":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "L"),
                "price": cols(["O", "P"]),
                "notes": cols(["M", "N", "Q", "R", "S"])
            }
        }

    return None

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    """
    Standardize COMPLANET price list.
    
    Args:
        input_path: Path to the Excel file
        output_path: Path for the output CSV
        include_no_price: If True, includes products even without prices (default: True)
    """
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'with_price': 0,
        'without_price': 0,
        'skipped_hidden': 0,
        'skipped_no_name': 0
    }

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    stats['skipped_hidden'] += 1
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. BUILD NAME FIRST (to check if row has data)
            name_cols = indices['name']
            name_parts = [clean_text(get_val(i)) for i in name_cols if get_val(i)]
            final_name = " ".join(name_parts)
            
            # Skip rows without name
            if not final_name:
                continue
                
            # Skip header rows (containing keywords like "MODEL NAME", "ЦЕНА")
            if any(keyword in final_name.upper() for keyword in ['MODEL NAME', 'ЦЕНА', 'PRICE']):
                continue
            
            stats['skipped_no_name'] += 1 if not final_name else 0

            # 3. EXTRACT PRICE (but don't require it)
            price_cols = indices['price']
            raw_price = ""
            for p_idx in price_cols:
                p_val = get_val(p_idx)
                if p_val and p_val != "0":
                    raw_price = p_val
                    break 
            
            # Process price
            clean_price = "0"
            has_price = False
            
            if raw_price:
                try:
                    numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                    if numeric_str:
                        clean_price = str(int(float(numeric_str)))
                        has_price = True
                except:
                    if "call" in raw_price.lower():
                        clean_price = "CALL"
                        has_price = True
                    else:
                        clean_price = "0"
            
            # Track statistics
            if has_price and clean_price != "0":
                stats['with_price'] += 1
            else:
                stats['without_price'] += 1
            
            # Skip if no price and include_no_price is False
            if not include_no_price and (not has_price or clean_price == "0"):
                continue

            # 4. BUILD BRAND
            brand_cols = indices['brand']
            brand_parts = [clean_text(get_val(i)) for i in brand_cols if get_val(i)]
            final_brand = " ".join(brand_parts)

            # 5. BUILD NOTES
            notes_cols = indices.get('notes', [])
            notes_parts = [clean_text(get_val(i)) for i in notes_cols if get_val(i)]
            final_notes = ", ".join(notes_parts)
            
            # Add note if no price
            if not has_price or clean_price == "0":
                if final_notes:
                    final_notes = "No price listed, " + final_notes
                else:
                    final_notes = "No price listed"

            products.append({
                "Supplier": "COMPLANET",
                "Category": clean_text(sheet_name),
                "Brand": final_brand, 
                "Model": "",
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD", 
                "Stock": "1",
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for COMPLANET.")
        return

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    
    # Print statistics
    print(f"Success! Converted {len(df_out)} items from COMPLANET.")
    print(f"  - Products with price: {stats['with_price']}")
    print(f"  - Products without price: {stats['without_price']}")
    if stats['skipped_hidden'] > 0:
        print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")
