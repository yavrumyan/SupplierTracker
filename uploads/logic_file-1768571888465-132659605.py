"""
conversion_logic_yt.py
Standardizes YT's price-lists (YT.xlsx).

Logic:
- Sheet Name = Category.
- IGNORES hidden rows.
- Fixes line breaks (\n) in cells.
- Header is Row 2.
- Layout A: Name = Col 1 + 2, Price = Col 3
- Layout B: Name = Col 1, Price = Col 2
"""

import pandas as pd
import openpyxl

def clean_text(text):
    """
    Safe helper to remove line breaks and extra spaces.
    Uses split() to handle all whitespace (newlines, tabs) automatically.
    """
    if not text:
        return ""
    # split() splits by ANY whitespace (space, tab, enter) and removes empty parts
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns the configuration for a given sheet name.
    Indices are 0-based.
    Header is Row 2 (so data starts at Row 3).
    """
    s_name = sheet_name.strip()

    # Layout A: Name = Cols 1+2 (Indices 0, 1), Price = Col 3 (Index 2)
    # Sheets: "LAN and Optical cable", "Server Cabinet"
    if s_name in ["LAN and Optical cable", "Server Cabinet"]:
        return {
            "header_row": 2,
            "indices": {
                "name": [0, 1], 
                "price": 2
            }
        }

    # Layout B: Name = Col 1 (Index 0), Price = Col 2 (Index 1)
    # Sheets: "POE SWICH,Convertor", "SFP", "UTP  Network ", "FO NETWORK 1", "FO NETWORK 2"
    # Note: We use flexible matching for "UTP Network" in case spaces vary
    elif s_name in ["POE SWICH,Convertor", "SFP", "FO NETWORK 1", "FO NETWORK 2"] or "UTP" in s_name:
        return {
            "header_row": 2,
            "indices": {
                "name": [0], 
                "price": 1
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load Workbook with openpyxl (needed for hidden rows check)
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        # Iterate Rows
        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SAFETY: Skip empty rows
            if not row:
                continue

            # 2. IGNORE HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 3. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                # Remove non-digits
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            # 4. BUILD NAME (Clean Newlines)
            name_parts = [clean_text(get_val(i)) for i in indices['name'] if get_val(i)]
            final_name = " ".join(name_parts)
            
            if not final_name:
                continue

            products.append({
                "Supplier": "YT",
                "Category": clean_text(sheet_name),
                "Brand": "", 
                "Model": "",
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD",
                "Stock": "1", # Default to 1
                "MOQ": "NO",
                "Notes": ""
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for YT.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from YT.")