"""
conversion_logic_complanet.py
Standardizes COMPLANET's price-lists.
Logic:
- IGNORES Hidden Rows.
- IGNORES Sheets not in the Allowlist.
- Fixes line breaks (\n).
- Default Stock = 1.
- Currency = USD.
- Keeps products with empty/zero prices (Notes added).
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """
    Converts Excel column letter to 0-based index.
    A -> 0, B -> 1, Z -> 25, AA -> 26, etc.
    """
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def get_sheet_config(sheet_name):
    s_name = sheet_name.strip()
    
    # Helper to generate range of indices from start col to end col (inclusive)
    def col_range(start_col, end_col):
        start_idx = col_index(start_col)
        end_idx = col_index(end_col)
        return list(range(start_idx, end_idx + 1))

    # Helper for specific list of cols
    def cols(col_list):
        return [col_index(c) for c in col_list]

    # --- CONFIGURATION ---
    
    if s_name == "NOTEBOOKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),
                "notes": cols(["AB", "AC", "AF", "AG"])
            }
        }
        
    elif s_name == "MONITORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "ALL IN ONE PC":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD"])
            }
        }

    elif s_name == "RAM":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "CPU":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "AA"),
                "price": cols(["AD", "AE"]),
                "notes": cols(["AB", "AC", "AF", "AG", "AH"])
            }
        }

    elif s_name == "HDD, SSD":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "CASE":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "MOTHER BOARDS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "X"),
                "price": cols(["AA", "AB"]),
                "notes": cols(["Y", "Z", "AC", "AD", "AE"])
            }
        }

    elif s_name == "COOLER  FAN": # Note double space
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "POWER SYPLY":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "VGA":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "UPS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["W", "X"]),
                "notes": cols(["U", "V", "Y", "Z", "AA"])
            }
        }

    elif s_name == "PRINTERS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "W"),
                "price": cols(["Z", "AA"]),
                "notes": cols(["X", "Y", "AB", "AC", "AD"])
            }
        }

    elif s_name == "DVD-RW":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"]) 
            }
        }

    elif s_name == "PROJECTORS":
        return {
            "header_row": 5,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NETWORKS":
        return {
            "header_row": 2,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "T"),
                "price": cols(["V", "W"]),
                "notes": cols(["T", "U", "X", "Y", "Z"])
            }
        }

    elif s_name == "USB FLASH":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "N"),
                "price": cols(["Q", "R"]),
                "notes": cols(["O", "P", "S", "T", "U"])
            }
        }

    elif s_name == "PC COMPOMEMTS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "U"),
                "price": cols(["X", "Y"]),
                "notes": cols(["V", "W", "Z", "AA", "AB"])
            }
        }

    elif s_name == "NOTEBOOK BAGS":
        return {
            "header_row": 3,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "R"),
                "price": cols(["U", "V"]),
                "notes": cols(["S", "T", "W", "X", "Y"])
            }
        }

    elif s_name == "SECURITY CAMERAS":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "V"),
                "price": cols(["Y", "Z"]),
                "notes": cols(["W", "X", "AA", "AB", "AC"])
            }
        }

    elif s_name == "Home appliances":
        return {
            "header_row": 4,
            "indices": {
                "brand": cols(["A", "B"]),
                "name": col_range("C", "L"),
                "price": cols(["O", "P"]),
                "notes": cols(["M", "N", "Q", "R", "S"])
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""