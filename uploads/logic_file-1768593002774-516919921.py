"""
conversion_logic_elcore.py
Standardizes ELCORE's price-lists.
ROBUST VERSION: Handles .xlsx, .xls, and handles "Empty File" errors gracefully.
"""

import pandas as pd
import openpyxl
import os

def clean_text(text):
    if not text:
        return ""
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    s_name = sheet_name.strip()

    # Ignored Sheets
    if s_name in ["Main", "Contacts", "Сервисные центры", "Полезные ссылки"]:
        return None

    # 1. Sheet: "Ноутбуки"
    if s_name == "Ноутбуки":
        return {"header_row": 1, "indices": {"model": 0, "brand": 1, "name": list(range(2, 13)), "price": 14, "notes": [15], "stock": 16}}

    # 2. Sheet: "Моноблоки и Копьютеры"
    elif s_name == "Моноблоки и Копьютеры":
        return {"header_row": 3, "indices": {"model": 0, "brand": 1, "name": list(range(2, 15)), "price": 16, "notes": [17], "stock": 18}}

    # 3. Sheet: "Печатная техника"
    elif s_name == "Печатная техника":
        return {"header_row": 1, "indices": {"model": 0, "brand": None, "name": list(range(1, 10)), "price": 11, "notes": [13], "stock": 12}}

    # 4. Sheet: "Мониторы"
    elif s_name == "Мониторы":
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1, 2, 3, 4, 5, 6, 7, 13, 14, 17, 18], "price": 20, "notes": [21], "stock": 22}}

    # 5. Sheet: "Планшеты"
    elif s_name == "Планшеты":
        return {"header_row": 4, "indices": {"model": 0, "brand": 1, "name": list(range(2, 9)), "price": 10, "notes": [11], "stock": 12}}

    # 6. Sheet: "Интерактивные дисплеи"
    elif s_name == "Интерактивные дисплеи":
        return {"header_row": 3, "indices": {"model": 1, "brand": None, "name": [0, 2], "price": 5, "notes": [6], "stock": 3}}

    # 7. Sheet: "UPS "
    elif "UPS" in s_name:
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1, 2, 3, 6, 7, 8], "price": 12, "notes": [14, 15], "stock": 13}}

    # 8. Sheet: "Xiaomi_ECO"
    elif s_name == "Xiaomi_ECO":
        return {"header_row": 3, "indices": {"model": 2, "brand": 0, "name": [1, 3], "price": 5, "notes": [6], "stock": 7}}

    # 9. Sheet: "Смартфоны"
    elif s_name == "Смартфоны":
        return {"header_row": 3, "indices": {"model": 0, "brand": 1, "name": list(range(2, 7)), "price": 8, "notes": [9], "stock": 10}}

    # 10. Sheet: "OEM"
    elif s_name == "OEM":
        return {"header_row": 3, "indices": {"model": 0, "brand": None, "name": [1], "price": 3, "notes": None, "stock": 4}}

    return None

def process_dataframe(df, sheet_name, config):
    """Processes a pandas DataFrame (used for Fallback mode)"""
    products = []
    header_idx = config['header_row'] - 1 # 0-based index
    
    # If file is short, skip
    if len(df) <= header_idx:
        return []
        
    # Slice dataframe to start after header
    # We don't reset index so we can match original rows roughly if needed
    df_data = df.iloc[header_idx+1:]
    indices = config['indices']

    for _, row in df_data.iterrows():
        # Helper to get value by integer index
        def get_val(idx):
            if idx is None: return ""
            if 0 <= idx < len(row):
                val = row.iloc[idx]
                return str(val).strip() if pd.notna(val) else ""
            return ""

        # Extract Price
        raw_price = get_val(indices['price'])
        if not raw_price: continue

        clean_price = "0"
        try:
            numeric_str = "".join(c for c in raw_price if c.isdigit())
            if numeric_str: clean_price = str(int(numeric_str))
        except: clean_price = "0"

        if clean_price == "0" and "call" not in raw_price.lower(): pass

        # Build Name
        name_idx = indices['name']
        if isinstance(name_idx, list):
            name_parts = [clean_text(get_val(i)) for i in name_idx if get_val(i)]
            final_name = " ".join(name_parts)
        else:
            final_name = clean_text(get_val(name_idx))
        if not final_name: continue

        # Build Notes
        notes_idx = indices.get('notes')
        final_notes = ""
        if notes_idx:
            if isinstance(notes_idx, list):
                final_notes = ", ".join([clean_text(get_val(i)) for i in notes_idx if get_val(i)])
            else:
                final_notes = clean_text(get_val(notes_idx))

        # Build Stock
        raw_stock = get_val(indices.get('stock'))
        clean_stock = "0"
        if raw_stock:
            try: clean_stock = str(int(float(raw_stock)))
            except: 
                if raw_stock != "0": clean_stock = "1"

        products.append({
            "Supplier": "ELCORE",
            "Category": clean_text(sheet_name),
            "Brand": clean_text(get_val(indices.get('brand'))),
            "Model": clean_text(get_val(indices['model'])),
            "Name": final_name,
            "Price": clean_price,
            "Currency": "AMD",
            "Stock": clean_stock,
            "MOQ": "NO",
            "Notes": final_notes
        })
    return products

def standardize(input_path: str, output_path: str):
    all_products = []
    
    # METHOD 1: Try OpenPyXL (Best for hidden rows)
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        # If successful, process using OpenPyXL
        for sheet_name in wb.sheetnames:
            config = get_sheet_config(sheet_name)
            if not config: continue
            
            ws = wb[sheet_name]
            start_row = config['header_row'] + 1
            indices = config['indices']

            for row in ws.iter_rows(min_row=start_row, values_only=False):
                if not row: continue
                # Check Hidden
                try:
                    if ws.row_dimensions[row[0].row].hidden: continue
                except: pass

                # Extract via same logic...
                cells = [cell.value for cell in row]
                def get_val(idx):
                    if idx is None: return ""
                    if 0 <= idx < len(cells):
                        val = cells[idx]
                        return str(val).strip() if val is not None else ""
                    return ""
                
                # (Logic Duplicated for OpenPyXL specifics)
                raw_price = get_val(indices['price'])
                if not raw_price: continue
                
                clean_price = "0"
                try:
                    numeric_str = "".join(c for c in raw_price if c.isdigit())
                    if numeric_str: clean_price = str(int(numeric_str))
                except: clean_price = "0"

                name_idx = indices['name']
                if isinstance(name_idx, list):
                    final_name = " ".join([clean_text(get_val(i)) for i in name_idx if get_val(i)])
                else:
                    final_name = clean_text(get_val(name_idx))
                if not final_name: continue

                notes_idx = indices.get('notes')
                final_notes = ""
                if notes_idx:
                    if isinstance(notes_idx, list):
                        final_notes = ", ".join([clean_text(get_val(i)) for i in notes_idx if get_val(i)])
                    else:
                        final_notes = clean_text(get_val(notes_idx))

                raw_stock = get_val(indices.get('stock'))
                clean_stock = "0"
                if raw_stock:
                    try: clean_stock = str(int(float(raw_stock)))
                    except: 
                        if raw_stock != "0": clean_stock = "1"
                
                all_products.append({
                    "Supplier": "ELCORE",
                    "Category": clean_text(sheet_name),
                    "Brand": clean_text(get_val(indices.get('brand'))),
                    "Model": clean_text(get_val(indices['model'])),
                    "Name": final_name,
                    "Price": clean_price,
                    "Currency": "AMD",
                    "Stock": clean_stock,
                    "MOQ": "NO",
                    "Notes": final_notes
                })

    except Exception as e:
        print(f"OpenPyXL failed ({e}). Falling back to Pandas...")
        # METHOD 2: Fallback to Pandas (Handles .xls, but ignores hidden rows)
        try:
            # sheet_name=None reads ALL sheets
            xls_dict = pd.read_excel(input_path, sheet_name=None, header=None)
            for sheet_name, df in xls_dict.items():
                config = get_sheet_config(sheet_name)
                if config:
                    all_products.extend(process_dataframe(df, sheet_name, config))
        except Exception as pd_e:
            print(f"Critical Error: Failed to read file with both methods. {pd_e}")
            return

    # Output
    df_out = pd.DataFrame(all_products)
    if df_out.empty:
        print("Warning: No products found for ELCORE.")
        return

    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from ELCORE.")