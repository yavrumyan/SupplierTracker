"""
conversion_logic_kosatec.py
Standardizes KOSATEC's price-lists (Excel Version).
Logic:
- Input: .xlsx file.
- Fetches real-time EUR->USD rate using YOUR API (EUR Base).
- DYNAMIC COLUMN MAPPING.
- Converts Price (EUR) to USD.
- Rounds Price to 2 decimal places.
- IGNORES Hidden Rows.
- FILTERS OUT products with 0 Stock (Critical for performance).
"""

import pandas as pd
import openpyxl
import urllib.request
import json
import warnings
import ssl

# --- CONFIGURATION ---
# Your API URL (EUR Base)
API_URL = "https://v6.exchangerate-api.com/v6/d77020a8cf31bc39255c92a2/latest/EUR"
# Fallback just in case even this API is blocked
MANUAL_RATE = 1.20  
# ---------------------

# Suppress warnings
warnings.filterwarnings("ignore")

def get_eur_to_usd_rate():
    """
    Fetches rates from exchangerate-api.com (EUR Base).
    Returns the EUR->USD multiplier directly.
    """
    try:
        # Create a context that doesn't verify SSL certificates strictly 
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        with urllib.request.urlopen(API_URL, context=ctx, timeout=5) as url:
            data = json.loads(url.read().decode())
            
            # The API returns EUR base. Example: "USD": 1.08 (1 EUR = 1.08 USD)
            # We use this rate directly.
            usd_rate = float(data['conversion_rates']['USD'])
            
            print(f"Fetched Rate (1 EUR = {usd_rate} USD)")
            return usd_rate

    except Exception as e:
        print(f"Notice: API blocked or failed ({e}). Using Manual Rate: {MANUAL_RATE}")
        return MANUAL_RATE

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def find_column_indices(ws, header_row_idx):
    mapping = {}
    keywords = {
        'model': ['mfrnr', 'model', 'part number'],
        'name': ['artname', 'name', 'description', 'bezeichnung'],
        'brand': ['manufacturer', 'brand', 'hersteller'],
        'price': ['cost', 'price', 'ek', 'preis'],
        'stock': ['quantity', 'stock', 'bestand', 'menge'],
        'notes': ['eta', 'note', 'remark']
    }
    
    for cell in ws[header_row_idx]:
        val = str(cell.value).lower().strip() if cell.value else ""
        if not val: continue
        
        for key, possible_names in keywords.items():
            if key not in mapping:
                for name in possible_names:
                    if name in val:
                        mapping[key] = cell.col_idx - 1
                        break
    return mapping

def standardize(input_path: str, output_path: str):
    print("Starting KOSATEC conversion (EUR Base API + Zero-Stock Filter)...")
    
    exchange_rate = get_eur_to_usd_rate()
    
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    skipped_zero_stock = 0
    ws = wb.active
    
    # Find Headers
    header_row = 1
    cols = {}
    for r in range(1, 6):
        found = find_column_indices(ws, r)
        if 'model' in found and 'price' in found:
            cols = found
            header_row = r
            print(f"Found headers on Row {r}: {cols}")
            break
    
    if not cols:
        print("Error: Could not detect standard Kosatec headers. Using fixed fallback.")
        cols = {'model': 1, 'name': 2, 'brand': 3, 'price': 6, 'stock': 9, 'notes': 10}
        header_row = 1

    start_row = header_row + 1
    
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        try:
            if ws.row_dimensions[row[0].row].hidden:
                continue
        except:
            pass

        cells = [cell.value for cell in row]
        
        def get_val(key):
            idx = cols.get(key)
            if idx is None: return ""
            if 0 <= idx < len(cells):
                val = cells[idx]
                return str(val).strip() if val is not None else ""
            return ""

        # STOCK (Check First)
        raw_stock = get_val('stock')
        clean_stock = "0"
        try:
            clean_str = "".join(c for c in raw_stock if c.isdigit() or c == '.')
            if clean_str:
                val = float(clean_str)
                clean_stock = str(int(val))
        except:
            clean_stock = "0"

        # FILTER: SKIP IF STOCK IS 0
        if clean_stock == "0":
            skipped_zero_stock += 1
            continue

        # NAME
        final_name = clean_text(get_val('name'))
        if not final_name:
            continue

        # PRICE
        raw_price = get_val('price')
        clean_price = "0"
        try:
            clean_str = raw_price.replace('€', '').strip()
            if ',' in clean_str:
                if '.' in clean_str and clean_str.find('.') < clean_str.find(','):
                    clean_str = clean_str.replace('.', '')
                    clean_str = clean_str.replace(',', '.')
                elif '.' in clean_str and clean_str.find(',') < clean_str.find('.'):
                    clean_str = clean_str.replace(',', '')
                else:
                    clean_str = clean_str.replace(',', '.')
            
            clean_str = "".join(c for c in clean_str if c.isdigit() or c == '.')
            
            if clean_str:
                val_eur = float(clean_str)
                val_usd = val_eur * exchange_rate
                val_usd = round(val_usd, 2)
                
                if val_usd.is_integer():
                    clean_price = str(int(val_usd))
                else:
                    clean_price = str(val_usd)
        except:
            clean_price = "0"

        # OTHER FIELDS
        final_model = clean_text(get_val('model'))
        final_brand = clean_text(get_val('brand'))
        
        # Notes
        final_notes = clean_text(get_val('notes'))
        # Fallback to col K (index 10) if dynamic map fails
        if not final_notes and len(cells) > 10:
             final_notes = clean_text(str(cells[10]))

        if final_notes.lower() == "nan": final_notes = ""

        products.append({
            "Supplier": "KOSATEC",
            "Category": "General", 
            "Brand": final_brand,
            "Model": final_model,
            "Name": final_name,
            "Price": clean_price,
            "Currency": "USD", 
            "Stock": clean_stock,
            "MOQ": "NO",
            "Notes": final_notes
        })

    # Output
    df_out = pd.DataFrame(products)
    
    if df_out.empty:
        print("Warning: No products found (or all had 0 stock).")
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from KOSATEC.")
    print(f"Skipped {skipped_zero_stock} items with 0 stock.")