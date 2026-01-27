"""
conversion_logic_muk.py
Standardizes MUK's price-lists.
Logic:
- Sheet name = Category
- IGNORES Hidden Rows
- Fixes line breaks (\n)
- Ignores all sheets except: Laptops, All in One, Monitors, Access
- Default Stock = 1 (unless specified in Status column)
- Currency = USD
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    """Remove extra whitespace and line breaks from text"""
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def col_index(col_str):
    """
    Converts Excel column letter to 0-based index.
    A -> 0, B -> 1, Z -> 25, AA -> 26, etc.
    """
    col_str = col_str.upper().strip()
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def get_sheet_config(sheet_name):
    """
    Returns configuration for each sheet.
    Returns None if sheet should be ignored.
    """
    s_name = sheet_name.strip()
    
    if s_name == "Laptops":
        # Header = row 1
        # Column A = Stock
        # Column C = Model
        # Column B+P = Name (columns 1 and 15)
        # Column D = Price (USD)
        return {
            "header_row": 1,
            "stock_col": col_index("A"),
            "model_col": col_index("C"),
            "name_cols": [col_index("B"), col_index("P")],
            "price_col": col_index("D")
        }
    
    elif s_name == "All in One":
        # Header = row 1
        # Column A+O = Name (columns 0 and 14)
        # Column B = Model
        # Column C = Price (USD)
        return {
            "header_row": 1,
            "stock_col": None,  # No stock column
            "model_col": col_index("B"),
            "name_cols": [col_index("A"), col_index("O")],
            "price_col": col_index("C")
        }
    
    elif s_name == "Monitors":
        # Header = row 1
        # Column A = Stock
        # Column C = Model
        # Column B+V = Name (columns 1 and 21)
        # Column D = Price (USD)
        return {
            "header_row": 1,
            "stock_col": col_index("A"),
            "model_col": col_index("C"),
            "name_cols": [col_index("B"), col_index("V")],
            "price_col": col_index("D")
        }
    
    elif s_name == "Access":
        # Header = row 1
        # Column A+D = Name (columns 0 and 3)
        # Column B = Model
        # Column C = Price (USD)
        return {
            "header_row": 1,
            "stock_col": None,  # No stock column
            "model_col": col_index("B"),
            "name_cols": [col_index("A"), col_index("D")],
            "price_col": col_index("C")
        }
    
    return None  # Ignore all other sheets

def parse_stock(stock_text):
    """
    Parse stock status from text.
    Returns stock quantity as string.
    """
    if not stock_text or pd.isna(stock_text):
        return "1"  # Default
    
    stock_str = str(stock_text).strip().upper()
    
    # Check for "STOCK" or "IN STOCK"
    if "STOCK" in stock_str and "LEFT" not in stock_str:
        return "1"  # In stock but quantity not specified
    
    # Check for "X UNITS LEFT" or "X UNIT LEFT"
    if "UNIT" in stock_str and "LEFT" in stock_str:
        try:
            # Extract number before "UNIT"
            parts = stock_str.split("UNIT")[0].strip().split()
            if parts:
                num_str = parts[-1]
                # Try to convert to number
                return str(int(num_str))
        except:
            pass
        return "1"
    
    # Check for "TRANSIT"
    if "TRANSIT" in stock_str:
        return "0"  # Transit means not in stock yet
    
    # Try to parse as direct number
    try:
        return str(int(float(stock_str)))
    except:
        return "1"  # Default if can't parse

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    """
    Standardize MUK price list.
    
    Args:
        input_path: Path to the Excel file
        output_path: Path for the output CSV
        include_no_price: If True, includes products even without prices (default: True)
    """
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'with_price': 0,
        'without_price': 0,
        'skipped_hidden': 0
    }

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue  # Skip sheets not in our list

        ws = wb[sheet_name]
        header_row = config['header_row']
        start_row = header_row + 1

        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    stats['skipped_hidden'] += 1
                    continue
            except:
                pass

            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: 
                    return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. EXTRACT MODEL (required)
            model = clean_text(get_val(config['model_col']))
            if not model:
                continue  # Skip rows without model
            
            # Skip header rows
            if model.upper() in ['MODEL', 'МОДЕЛЬ']:
                continue

            # 3. EXTRACT PRICE
            raw_price = get_val(config['price_col'])
            clean_price = "0"
            has_price = False
            
            if raw_price:
                try:
                    # Remove any non-numeric characters except decimal point
                    numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                    if numeric_str:
                        price_float = float(numeric_str)
                        # Format to 2 decimal places, but remove trailing zeros
                        clean_price = f"{price_float:.2f}".rstrip('0').rstrip('.')
                        has_price = True
                except:
                    if "call" in raw_price.lower():
                        clean_price = "CALL"
                        has_price = True
                    else:
                        clean_price = "0"
            
            # Track statistics
            if has_price and clean_price not in ["0", "CALL"]:
                stats['with_price'] += 1
            else:
                stats['without_price'] += 1
            
            # Skip if no price and include_no_price is False
            if not include_no_price and (not has_price or clean_price == "0"):
                continue

            # 4. BUILD NAME from multiple columns
            name_cols = config['name_cols']
            name_parts = [clean_text(get_val(col_idx)) for col_idx in name_cols if get_val(col_idx)]
            final_name = " ".join(name_parts)
            
            if not final_name:
                final_name = model  # Use model as fallback name
            
            # 5. EXTRACT STOCK
            stock_col = config.get('stock_col')
            if stock_col is not None:
                raw_stock = get_val(stock_col)
                final_stock = parse_stock(raw_stock)
            else:
                final_stock = "1"  # Default stock for sheets without stock column

            # 6. CREATE NOTES (empty for MUK, but can add stock status if needed)
            notes = ""
            if not has_price or clean_price == "0":
                notes = "No price listed"

            products.append({
                "Supplier": "MUK",
                "Category": clean_text(sheet_name),
                "Brand": "",  # MUK doesn't have brand column
                "Model": model,
                "Name": final_name,
                "Price": clean_price,
                "Currency": "USD",
                "Stock": final_stock,
                "MOQ": "NO",
                "Notes": notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for MUK.")
        # Create empty template
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])
    
    # Force string type for all columns
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    
    # Print statistics
    print(f"Success! Converted {len(df_out)} items from MUK.")
    print(f"  - Products with price: {stats['with_price']}")
    print(f"  - Products without price: {stats['without_price']}")
    if stats['skipped_hidden'] > 0:
        print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")
