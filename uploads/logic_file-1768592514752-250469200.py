"""
conversion_logic_elcore.py
Standardizes ELCORE's price-lists.

Logic:
- Sheet Name = Category.
- IGNORES Hidden Rows.
- Fixes line breaks (\n).
- Ignores specific sheets (Main, Contacts, etc.).
- Complex column merging for Name/Notes per sheet.
"""

import pandas as pd
import openpyxl

def clean_text(text):
    """
    Safe helper to remove line breaks and extra spaces.
    """
    if not text:
        return ""
    # split() handles all whitespace (newlines, tabs) automatically
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns configuration for a given sheet.
    header_row is 1-based Excel row index.
    indices are 0-based (Col A=0).
    """
    s_name = sheet_name.strip()

    # Ignored Sheets
    if s_name in ["Main", "Contacts", "Сервисные центры", "Полезные ссылки"]:
        return None

    # 1. Sheet: "Ноутбуки"
    # Header 1st row (Row 1). Data starts Row 2.
    if s_name == "Ноутбуки":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": 1, 
                "name": list(range(2, 13)), # Cols 3 to 13 (Indices 2-12)
                "price": 14, "notes": [15], "stock": 16
            }
        }

    # 2. Sheet: "Моноблоки и Копьютеры"
    # Header 3rd row. Data starts Row 4.
    elif s_name == "Моноблоки и Копьютеры":
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": 1, 
                "name": list(range(2, 15)), # Cols 3 to 15 (Indices 2-14)
                "price": 16, "notes": [17], "stock": 18
            }
        }

    # 3. Sheet: "Печатная техника"
    # Header 1st row.
    elif s_name == "Печатная техника":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": None, 
                "name": list(range(1, 10)), # Cols 2 to 10 (Indices 1-9)
                "price": 11, "notes": [13], "stock": 12
            }
        }

    # 4. Sheet: "Мониторы"
    # Header 3rd row.
    # Name indices: 1-7, 13, 14, 17, 18 (0-based)
    elif s_name == "Мониторы":
        name_cols = [1, 2, 3, 4, 5, 6, 7, 13, 14, 17, 18]
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": None, 
                "name": name_cols,
                "price": 20, "notes": [21], "stock": 22
            }
        }

    # 5. Sheet: "Планшеты"
    # Header 4th row.
    elif s_name == "Планшеты":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": 1, 
                "name": list(range(2, 9)), # Cols 3 to 9 (Indices 2-8)
                "price": 10, "notes": [11], "stock": 12
            }
        }

    # 6. Sheet: "Интерактивные дисплеи"
    # Header 3rd row.
    elif s_name == "Интерактивные дисплеи":
        return {
            "header_row": 3,
            "indices": {
                "model": 1, "brand": None, 
                "name": [0, 2], 
                "price": 5, "notes": [6], "stock": 3
            }
        }

    # 7. Sheet: "UPS " (Note space)
    # Header 3rd row.
    elif "UPS" in s_name:
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": None, 
                "name": [1, 2, 3, 6, 7, 8], 
                "price": 12, "notes": [14, 15], "stock": 13
            }
        }

    # 8. Sheet: "Xiaomi_ECO"
    # Header 3rd row.
    elif s_name == "Xiaomi_ECO":
        return {
            "header_row": 3,
            "indices": {
                "model": 2, "brand": 0, 
                "name": [1, 3], 
                "price": 5, "notes": [6], "stock": 7
            }
        }

    # 9. Sheet: "Смартфоны"
    # Header 3rd row.
    elif s_name == "Смартфоны":
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": 1, 
                "name": list(range(2, 7)), 
                "price": 8, "notes": [9], "stock": 10
            }
        }

    # 10. Sheet: "OEM"
    # Header 3rd row.
    elif s_name == "OEM":
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": None, 
                "name": [1], 
                "price": 3, "notes": None, "stock": 4
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load Workbook with data_only=True to get calculated values
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        # Iterate Rows
        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SAFETY: Skip empty rows
            if not row:
                continue

            # 2. IGNORE HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 3. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                # Remove non-digits
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass 

            # 4. BUILD NAME (Merge Cols + Clean Text)
            name_idx = indices['name']
            if isinstance(name_idx, list):
                name_parts = [clean_text(get_val(i)) for i in name_idx if get_val(i)]
                final_name = " ".join(name_parts)
            else:
                final_name = clean_text(get_val(name_idx))
            
            if not final_name:
                continue

            # 5. BUILD NOTES
            notes_idx = indices.get('notes')
            final_notes = ""
            if notes_idx:
                if isinstance(notes_idx, list):
                    note_parts = [clean_text(get_val(i)) for i in notes_idx if get_val(i)]
                    final_notes = ", ".join(note_parts)
                else:
                    final_notes = clean_text(get_val(notes_idx))

            # 6. STOCK
            raw_stock = get_val(indices.get('stock'))
            clean_stock = "0"
            # Try to interpret stock
            if raw_stock:
                try:
                    # If number, use it
                    clean_stock = str(int(float(raw_stock)))
                except:
                    # If text (e.g. "+", "Yes", "On Order"), treat as 1 if it implies availability
                    # For safety, let's just default text to "1" unless it says "0"
                    if raw_stock != "0":
                        clean_stock = "1"
            
            products.append({
                "Supplier": "ELCORE",
                "Category": clean_text(sheet_name),
                "Brand": clean_text(get_val(indices.get('brand'))),
                "Model": clean_text(get_val(indices['model'])),
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD",
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for ELCORE.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from ELCORE.")