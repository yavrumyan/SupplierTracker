"""
conversion_logic_xart.py
Standardizes X-ART's price-lists.
Logic:
- IGNORES Hidden Rows.
- IGNORES Sheets not in the Allowlist.
- Fixes line breaks (\n) in cells.
- Sheet-specific header rows and column mappings.
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    """
    Removes line breaks, tabs, and extra spaces.
    """
    if not text or pd.isna(text):
        return ""
    # split() splits by ANY whitespace (newline, tab, space) and joins with single space
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns configuration for a given sheet name.
    header_row: 1-based Excel row number.
    indices: 0-based column indices (Col A=0, B=1, etc).
    """
    s_name = sheet_name.strip()

    # --- QSAN ---
    # Header 9, Model(2), Brand(2), Name(3-11), Price(13), Notes(12), Stock(14)
    if s_name == "QSAN":
        return {
            "header_row": 9,
            "indices": {
                "model": 1, "brand": 1, 
                "name": list(range(2, 11)), # Cols 3-11 -> Indices 2-10
                "price": 12, "notes": [11], "stock": 13
            }
        }

    # --- Ugreen NAS ---
    # Header 2, Model(1), Name(3-10), Price(11), Notes(12)
    elif s_name == "Ugreen NAS":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": list(range(2, 10)), # Cols 3-10 -> Indices 2-9
                "price": 10, "notes": [11], "stock": None
            }
        }

    # --- Ugreen ---
    # Header 2, Model(1), Name(2), Price(3), Notes(4-7)
    elif s_name == "Ugreen":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5, 6], "stock": None
            }
        }

    # --- Ugreen mobile ---
    # Header 2, Model(1), Name(2), Price(3), Notes(4-7)
    elif s_name == "Ugreen mobile":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5, 6], "stock": None
            }
        }

    # --- Ugreen Network ---
    # Header 3, Model(1), Name(2), Price(3), Notes(4-7)
    elif s_name == "Ugreen Network":
        return {
            "header_row": 3,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5, 6], "stock": None
            }
        }

    # --- nJoy ---
    # Header 6, Model(1), Name(2), Price(3), Stock(6), Notes(4,5,7,8,9,10)
    elif s_name == "nJoy":
        return {
            "header_row": 6,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, 
                # Notes: 4,5 (Indices 3,4) and 7,8,9,10 (Indices 6,7,8,9). Skipping Col 6 (Index 5)
                "notes": [3, 4, 6, 7, 8, 9], 
                "stock": 5
            }
        }

    # --- Ajax Systems ---
    # Header 2, Name(2), Price(3), Notes(4-7), Stock(6)
    # Note: Col 6 is Stock, but user asked for 4+5+6+7 in Notes. Included both.
    elif s_name == "Ajax Systems":
        return {
            "header_row": 2,
            "indices": {
                "model": None, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5, 6], "stock": 5
            }
        }

    # --- Logitech ---
    # Header 5, Model(1), Name(2), Price(3), Notes(4-6)
    elif s_name == "Logitech":
        return {
            "header_row": 5,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5], "stock": None
            }
        }

    # --- Thermaltake ---
    # Header 2, Model(1), Name(2), Price(3), Notes(4-6)
    elif s_name == "Thermaltake":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4, 5], "stock": None
            }
        }

    # --- Cooler Master ---
    # Header 2, Model(1), Name(2), Price(5)
    elif s_name == "Cooler Master":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 4, "notes": None, "stock": None
            }
        }

    # --- CSB Battery ---
    # Header 2, Model(1), Name(2), Price(3)
    elif s_name == "CSB Battery":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": None, "stock": None
            }
        }

    # --- Hyperline ---
    # Header 5, Model(1), Name(2), Price(3), Notes(4)
    elif s_name == "Hyperline":
        return {
            "header_row": 5,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3], "stock": None
            }
        }

    # --- WRLine ---
    # Header 2, Model(1), Name(2), Price(3), Notes(4-5)
    elif s_name == "WRLine":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    # --- Draka ---
    # Header 4, Model(1), Name(2), Price(3), Notes(4-5)
    elif s_name == "Draka":
        return {
            "header_row": 4,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    # --- Kopos ---
    # Header 7, Model(1), Name(2), Price(3), Notes(4-5)
    elif s_name == "Kopos":
        return {
            "header_row": 7,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    # --- Ning Bo ---
    # Header 1, Model(1), Name(2), Price(3), Notes(4-5)
    elif s_name == "Ning Bo":
        return {
            "header_row": 1,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    # --- ZyXEL ---
    # Header 2, Model(1), Name(2), Price(3), Notes(4-5)
    elif s_name == "ZyXEL":
        return {
            "header_row": 2,
            "indices": {
                "model": 0, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    # --- netping ---
    # Header 2, Name(2), Price(3), Notes(4-5)
    elif s_name == "netping":
        return {
            "header_row": 2,
            "indices": {
                "model": None, "brand": None,
                "name": [1], 
                "price": 2, "notes": [3, 4], "stock": None
            }
        }

    return None

def standardize(input_path: str, output_path: str):
    try:
        # Load workbook using openpyxl to access hidden row properties
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []

    for sheet_name in wb.sheetnames:
        config = get_sheet_config(sheet_name)
        if not config:
            # Skip ignored sheets
            continue

        ws = wb[sheet_name]
        start_row = config['header_row'] + 1 
        indices = config['indices']

        # Iterate Rows
        for row in ws.iter_rows(min_row=start_row, values_only=False):
            # 1. SKIP HIDDEN ROWS
            try:
                if ws.row_dimensions[row[0].row].hidden:
                    continue
            except:
                pass

            # Extract cell values safely
            cells = [cell.value for cell in row]
            
            def get_val(idx):
                if idx is None: return ""
                if 0 <= idx < len(cells):
                    val = cells[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            # 2. EXTRACT PRICE
            raw_price = get_val(indices['price'])
            if not raw_price:
                continue 

            clean_price = "0"
            try:
                numeric_str = "".join(c for c in raw_price if c.isdigit())
                if numeric_str:
                    clean_price = str(int(numeric_str))
            except:
                clean_price = "0"

            if clean_price == "0" and "call" not in raw_price.lower():
                pass 

            # 3. BUILD NAME
            name_idx = indices['name']
            if isinstance(name_idx, list):
                name_parts = [clean_text(get_val(i)) for i in name_idx if get_val(i)]
                final_name = " ".join(name_parts)
            else:
                final_name = clean_text(get_val(name_idx))
            
            if not final_name:
                continue

            # 4. BUILD NOTES
            notes_idx = indices.get('notes')
            final_notes = ""
            if notes_idx:
                if isinstance(notes_idx, list):
                    final_notes = ", ".join([clean_text(get_val(i)) for i in notes_idx if get_val(i)])
                else:
                    final_notes = clean_text(get_val(notes_idx))

            # 5. STOCK
            # If stock config is present, use it. Otherwise default to "1".
            if indices.get('stock') is not None:
                raw_stock = get_val(indices['stock'])
                clean_stock = "0"
                if raw_stock:
                    try:
                        clean_stock = str(int(float(raw_stock)))
                    except:
                        if raw_stock != "0": clean_stock = "1"
            else:
                clean_stock = "1"
            
            products.append({
                "Supplier": "X-ART",
                "Category": clean_text(sheet_name),
                "Brand": clean_text(get_val(indices.get('brand'))),
                "Model": clean_text(get_val(indices.get('model'))),
                "Name": final_name,
                "Price": clean_price,
                "Currency": "AMD",
                "Stock": clean_stock,
                "MOQ": "NO",
                "Notes": final_notes
            })

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for X-ART.")
        return

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    # Save
    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from X-ART.")