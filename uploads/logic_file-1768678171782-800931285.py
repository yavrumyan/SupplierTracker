"""
conversion_logic_haifa.py
Standardizes HAIFA's price-lists.
Logic:
- TARGET: Only processes sheet "Full list".
- IGNORES Hidden Rows.
- Fixes line breaks.
- Header = Row 7.
- Columns: A=Model, C=Name, D=Stock, E=Price(USD), F=Notes.
"""

import pandas as pd
import openpyxl
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    return " ".join(str(text).split())

def get_sheet_config(sheet_name):
    """
    Returns config ONLY for 'Full list' sheet.
    Matches "Full list", "Full list ", etc.
    """
    s_name = sheet_name.strip()
    
    if s_name.lower() == "full list":
        return {
            "header_row": 7,  # Data starts at row 8
            "indices": {
                "model": 0,   # Col A
                "name": [2],  # Col C
                "stock": 3,   # Col D
                "price": 4,   # Col E
                "notes": [5]  # Col F
            }
        }
    
    return None

def standardize(input_path: str, output_path: str, include_no_price: bool = True):
    try:
        # Load workbook with openpyxl to access hidden rows
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    products = []
    stats = {
        'processed': 0,
        'skipped_hidden': 0
    }

    # Only look for the "Full list" sheet
    target_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() == "full list":
            target_sheet = name
            break
    
    if not target_sheet:
        print("Warning: Sheet 'Full list' not found in HAIFA.xlsx")
        return

    ws = wb[target_sheet]
    config = get_sheet_config(target_sheet)
    start_row = config['header_row'] + 1 
    indices = config['indices']

    for row in ws.iter_rows(min_row=start_row, values_only=False):
        # 1. SKIP HIDDEN ROWS
        try:
            if ws.row_dimensions[row[0].row].hidden:
                stats['skipped_hidden'] += 1
                continue
        except:
            pass

        cells = [cell.value for cell in row]
        
        def get_val(idx):
            if idx is None: return ""
            if 0 <= idx < len(cells):
                val = cells[idx]
                return str(val).strip() if val is not None else ""
            return ""

        # 2. NAME (Required)
        name_parts = [clean_text(get_val(i)) for i in indices['name'] if get_val(i)]
        final_name = " ".join(name_parts)
        if not final_name:
            continue

        # 3. PRICE
        raw_price = get_val(indices['price'])
        clean_price = "0"
        
        if raw_price:
            try:
                # Remove '$' and ',' then convert
                numeric_str = "".join(c for c in raw_price if c.isdigit() or c == '.')
                if numeric_str:
                    clean_price = str(int(float(numeric_str)))
            except:
                if "call" in raw_price.lower():
                    clean_price = "CALL"
                else:
                    clean_price = "0"

        # 4. STOCK
        raw_stock = get_val(indices['stock'])
        clean_stock = "0"
        if raw_stock:
            try:
                # Try to parse number (e.g. "5", "5.0")
                val = float(raw_stock)
                clean_stock = str(int(val))
            except:
                # If text (e.g., "+", "Yes"), assume 1 if not negative/zero words
                if raw_stock not in ["0", "-", "no"]:
                    clean_stock = "1"

        # 5. MODEL & NOTES
        final_model = clean_text(get_val(indices['model']))
        
        notes_parts = [clean_text(get_val(i)) for i in indices['notes'] if get_val(i)]
        if clean_price == "0":
            notes_parts.append("Price Not Specified")
        final_notes = ", ".join(notes_parts)

        products.append({
            "Supplier": "HAIFA",
            "Category": "Full list", # Using sheet name as category
            "Brand": "", 
            "Model": final_model,
            "Name": final_name,
            "Price": clean_price,
            "Currency": "USD", 
            "Stock": clean_stock,
            "MOQ": "NO",
            "Notes": final_notes
        })
        stats['processed'] += 1

    # Create DataFrame
    df_out = pd.DataFrame(products)

    if df_out.empty:
        print("Warning: No products found for HAIFA.")
        # Create empty template to prevent errors
        df_out = pd.DataFrame(columns=["Supplier","Category","Brand","Model","Name","Price","Currency","Stock","MOQ","Notes"])

    # JSON Safety
    for col in df_out.columns:
        df_out[col] = df_out[col].astype(str)

    df_out.to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"Success! Converted {len(df_out)} items from HAIFA.")
    print(f"  - Skipped hidden rows: {stats['skipped_hidden']}")